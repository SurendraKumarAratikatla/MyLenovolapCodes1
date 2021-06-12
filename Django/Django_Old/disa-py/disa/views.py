from django.shortcuts import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from member.models import Awardee, Seva, SevaCategory, SevaAddress, Organisation
from member.models import User
from member.models import *
from django.contrib.auth import login as auth_login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.views import logout as auth_logout
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.forms import ModelForm, Textarea, TextInput
from django.utils.translation import ugettext_lazy as _
from django import forms
from member.models import MaasamType, PakshamType, TithiType, RasiType
from django.db import models 
from datetime import datetime
from django.utils.encoding import smart_str
import csv
import uuid
import time
from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.views.generic import ListView
from django.utils.safestring import mark_safe
from django.forms.widgets import RadioSelect
from django.db import connection
from disa.middleware.current_user import get_current_user
from django.template.context_processors import request
from rolepermissions.decorators import has_role_decorator, has_permission_decorator
from rolepermissions.mixins import HasPermissionsMixin, HasRoleMixin
from django.utils.crypto import get_random_string
from django.contrib.auth import get_user_model
User = get_user_model()

class OrganisationForm(ModelForm):
    def __init__(self, *args, **kwargs):
        super(OrganisationForm, self).__init__(*args, **kwargs)
        self.fields['code'].label_css = "required"
        self.fields['name'].label_css = "required"
        self.fields['uri'].label_css = "required"
        empty_list = [("", "-----------")]
        empty_list = empty_list + list(Organisation.objects.values_list('id',
                                                                        'name'))
        self.fields['parentid'] = forms.ChoiceField(choices=empty_list,
                                                    required=False)

    class Meta:
        model = Organisation
        fields = ['code', 'uri', 'name', 'description', 'parentid', 'address', 'city','state', 'country', 'pincode', 'email', 'phone', 'mobile', 'fax', 'website']
        widgets = {
            'description': Textarea(attrs={'cols': 3, 'rows': 3, 'class': 'span4'}),
            'address': Textarea(attrs={'cols': 3, 'rows': 3, 'class': 'span4'}),
        }
        labels = {
            'code': _('Organisation code'),
            'uri':  _('Organisation uri'),
            'parentid': _("Parent Id")
        }
gender_options = ( ('m','Male'),
                   ('f','Female'),
                   ('d','Data Not Available')
                 )
class HorizRadioRenderer(forms.RadioSelect.renderer):
    """ this overrides widget method to put radio buttons horizontally
        instead of vertically.
    """
    def render(self):
            """Outputs radios"""
            return mark_safe(u'\n'.join([u'%s\n' % w for w in self]))


               
class MemberForm(ModelForm):
    def __init__(self, *args, **kwargs):
        super(MemberForm, self).__init__(*args, **kwargs)
        self.fields['place'] = forms.CharField(max_length=63, required=True)
        self.fields['name'] = forms.CharField(max_length=63, required=True)
        self.fields['surname'] = forms.CharField(max_length=63, required=True)
    
        self.fields['gender'].widget = RadioSelect(renderer=HorizRadioRenderer, choices=gender_options)
        self.fields['gender'].value='d'
        self.fields['calendar_type'].widget = RadioSelect(renderer=HorizRadioRenderer, choices=calendar_options)
        self.fields['calendar_type'].value='default'

    class Meta:
        model = Member
        fields = ['mid',
                  'id',
                  'created',
                  'updated',
                  'created_by',
                  'updated_by',
                  'salutation',
                  'surname',
                  'name',
                  'place',
                  'email',
                  'mobile',
                  'phone',
                  'calendar_type',
                  'maasam',
                  'paksham',
                  'tithi',
                  'rasi',
                  'birth_date',
                  'gotram',
                  'nakshatram',
                  'gender',
                  'dob'
                  ]


class AddressForm(ModelForm):
    class Meta:
        model = Address
        fields = ['id',
                  'created',
                  'updated',
                  'created_by',
                  'updated_by',
                  'mid',
                  'address',
                  'city',
                  'district',
                  'state',
                  'country',
                  'pin',
                  'is_valid',
                  'is_primary']


class HomeView(View):
    def get(self, request):
        today_date = time.strftime("%m%d")
        query_param = {'sevaday__icontains': today_date}
        seva_list = Seva.objects.filter(**query_param)
        return render(request, "home.html", locals())


class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated():
            return HttpResponseRedirect(reverse('home'))
        form = AuthenticationForm()
        return render(request, 'login.html', locals())

    def post(self, request):
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            
            user = User.objects.filter(pk=form.get_user_id())
            loginuser = form.get_user()
            if loginuser is not None:
                if loginuser.is_active:
                    auth_login(request, loginuser)
            return HttpResponseRedirect(reverse('home'))
        return render(request, 'login.html', locals())


class LogoutView(View):
    def get(self, request):
        auth_logout(request)
        return HttpResponseRedirect(reverse('landing'))


@method_decorator(login_required, name="get")
class OrganisationView(HasPermissionsMixin, View):

    required_permission = 'VIEW_ORGANISATION'
    
    def get(self, request):
        org_list = Organisation.objects.all()
        return render(request, 'organisation.html', locals())

@method_decorator(login_required, name="get")
class UserProfileView(HasPermissionsMixin, View):
    required_permission = 'VIEW_SELF_DETAILS'
    def get(self, request, mid=None):
        user = User.objects.get(pk=get_current_user().id)
        mid = user.mid
        if mid == None:
            return render(request, 'home.html', locals())    
        mem = Member.objects.get(pk=mid)
    
        if Address.objects.filter(mid=mid):
            address_list = Address.objects.filter(mid=mid)
        if Seva.objects.filter(mid=mem.id):
            sevas = Seva.objects.filter(mid=mem.id)
            sevaCategories = SevaCategory.objects.all()
        return render(request, 'member_view.html', locals())

@method_decorator(login_required, name="get")
class MemberView(HasPermissionsMixin, View):
    required_permission = 'VIEW_MEMBER_DETAILS'
    
    def get(self, request, mid=None):
        sevaaddress_list = []
        trustee = None
        staffprofile = None
        honorary = None
        mem = Member.objects.get(pk=mid)
        assetBuildings = AssetBuilding.objects.filter(maindonors=mid)
        assetLands = AssetLand.objects.filter(inchargeperson=mid)
        assetMoney =  DonationCash.objects.filter(member=mid)
        assetEquipments = AssetEquipment.objects.filter(maindonors=mid)
        try:
            trustee = Trustee.objects.get(member=mid)
        except:
            pass
        try:
            staffprofile = StaffProfile.objects.get(mid=mid)
        except:
            pass
        try:
            honorary = Honorary.objects.get(member=mid)
        except:
            pass

        if Address.objects.filter(mid=mid):
            address_list = Address.objects.filter(mid=mid)
        if Seva.objects.filter(mid=mem.id):
            sevas = Seva.objects.filter(mid=mem.id)
            for seva in sevas:
                if seva.inthenameof != 'Donor' and seva.inthenameof != 'donor':
                    try:
                        sevaaddress = SevaAddress.objects.get(sid=seva.id)
                        sevaaddress_list.append(sevaaddress)
                    except:
                        pass
            sevaCategories = SevaCategory.objects.all()
        return render(request, 'member_view.html', locals())


@method_decorator(login_required, name="get")
class MSearchView(HasPermissionsMixin, View):
    required_permission = 'VIEW_MEMBER_DETAILS'
   
    def get(self, request):
        search_by = request.GET.get("q", None)
        data = request.GET
        query_format = {}
        if search_by == "advanced":
            if data.get('name'):
                query_format['name__icontains'] = data["name"]
            if data.get('place'):
                query_format['place__icontains'] = data["place"]
            if data.get('email'):
                query_format['email__icontains'] = data["email"]
            if data.get('surname'):
                query_format['surname__icontains'] = data["surname"]
            if data.get('mobile'):
                query_format['mobile__icontains'] = data["mobile"]
            if data.get('phone'):
                query_format['phone__icontains'] = data["phone"]
        elif search_by == "mid":
            if data.get('mid'):
                query_format['mid__exact'] = data["mid"]
        if query_format:
            result = Member.objects.filter(**query_format)
        return render(request, 'msearch.html', locals())

@method_decorator(login_required, name="get")
class SSearchView(HasPermissionsMixin, View):
    required_permission = 'EXPORT_REPORTS'
   
    def get(self, request):
        export = request.GET.get("export", None)
        sid = request.GET.get("sid", None)
        from_date = request.GET.get("from", None)
        to_date = request.GET.get("to", None)
        query_param = {'sid': sid}
        if from_date and to_date:
            from_date = "".join(from_date.split("-")[1:])
            to_date = "".join(to_date.split("-")[1:])
            query_param = {'sid': sid, 'sevaday__range': [from_date, to_date]}
        seva_list = Seva.objects.order_by('sevaday').filter(**query_param)
        return render(request, 'ssearch.html', locals())

@method_decorator(login_required, name="get")
class ReportSevaTemplates(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    
    model = SevaCategory
    context_object_name = 'sevas'
    template_name = "reports_seva_templates.html"


@method_decorator(login_required, name="get")
class ReportsSeva(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = Organisation
    context_object_name = 'organisation'
    template_name = "reports_sevas.html"

    def get_queryset(self):
        return Organisation.objects.all()


@method_decorator(login_required, name="get")
class SevaCategoryView(HasPermissionsMixin, ListView):
    required_permission = 'VIEW_SEVA_CATEGORY'
    model = SevaCategory
    context_object_name = 'seva_cat_list'
    template_name = "seva_cat.html"


@method_decorator(login_required, name="get")
class SevaCategorySearch(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = SevaCategory
    context_object_name = 'seva_cat_list'
    template_name = "scatsearch.html"

    def get_queryset(self):
        oid = self.request.GET.get("oid", None)
        return SevaCategory.objects.filter(oid=oid)


@method_decorator(login_required, name="get")
class LocationSearch(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = Member
    template_name = "lsearch.html"

    def get_context_data(self, **kwargs):
        context = super(LocationSearch, self).get_context_data(**kwargs)
        loc = self.request.GET.get("loc", None)
        context['loc'] = loc
        context['loc_list'] = Member.objects.filter(place__icontains=loc)
        return context


@method_decorator(login_required, name="get")
class ReportsMonthly(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = SevaCategory
    context_object_name = 'sevas'
    template_name = "reports_monthly.html"


@method_decorator(login_required, name="get")
class SevaMonthlySearch(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = Seva
    template_name = "ssearch.html"

    def get_context_data(self, **kwargs):
        context = super(SevaMonthlySearch, self).get_context_data(**kwargs)
        sid = self.request.GET.get("sid", None)
        month_input = self.request.GET.get("month", None)
        context['sid'] = sid
        context['month_input'] = month_input
        query_param = {'sid': sid, 'sevadate__icontains': "%s%s%s" % ("-", month_input, "-")}
        context['seva_list'] = Seva.objects.filter(**query_param)
        return context


@method_decorator(login_required, name="get")
class ReportsMaasam(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    sevas = SevaCategory.objects.all()
    maasams = MaasamType.objects.all()
    template_name = "reports_maasam.html"

@method_decorator(login_required, name="get")
class SevaMaasamSearch(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = Seva 
    template_name = "lunarssearch.html"
 
    def get_context_data(self, **kwargs):
        context = super(SevaMaasamSearch, self).get_context_data(**kwargs)
        sid = self.request.GET.get("sid", None)
        maasam_input = self.request.GET.get("maasam", None)
        context['sid'] = sid
        context['maasam_input'] = maasam_input
        query_param = {'sid': sid, 'maasam':maasam_input}
        context['seva_list'] = Seva.objects.filter(**query_param)
        return context

@method_decorator(login_required, name="get")
class ReportsAwardee(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = Awardee
    context_object_name = 'awardees'
    template_name = "reports_awardee.html"

@method_decorator(login_required, name="get")
class AwardeeSearch(HasPermissionsMixin, ListView):
    required_permission = 'EXPORT_REPORTS'
    model = Awardee 
    context_object_name = 'awardees'
    template_name = "awardee_search.html"
 
    def get_context_data(self, **kwargs):
        context = super(AwardeeSearch, self).get_context_data(**kwargs)
        search_by = self.request.GET.get("q", None)
        data = self.request.GET
        query_format = {}
        if search_by == "year":
            if data.get('year'):
                query_format['year__icontains'] = data["year"]
                searchkey = data["year"]
        elif search_by == "specialization":
            if data.get('specialization'):
                query_format['specialization__icontains'] = data["specialization"]
                searchkey = data["specialization"]
        elif search_by == "name":
            if data.get('name'):
                query_format['name__icontains'] = data["name"]
                searchkey = data["name"]
        
        if query_format:
            context['awardees'] = Awardee.objects.filter(**query_format)
        context['searchkey'] = searchkey
        context['search_by'] = search_by
        return context

@method_decorator(login_required, name="get")
class MemberEditView(View):
    @method_decorator(has_permission_decorator('EDIT_MEMBER_DETAILS'))
    def get(self, request, m_id):
        data = Member.objects.get(id=m_id)
        form = MemberForm(instance=data)
        maasam = MaasamType.objects.all()
        paksham = PakshamType.objects.all()
        tithi = TithiType.objects.all()
        rasi = RasiType.objects.all()
        nakshatrams = NakshatramType.objects.all()
        return render(request, 'edit_member.html', locals())

    @method_decorator(has_permission_decorator('EDIT_MEMBER_DETAILS'))
    def post(self, request, m_id):
        data = Member.objects.get(id=m_id)
        data1 = dict(data.__dict__.items() + request.POST.items())
        data1['updated'] = datetime.now()
        data1['updated_by'] = get_current_user().username
        dob =  request.POST.get('dob')

        if data.calendar_type != data1['calendar_type']:
            if data.calendar_type == 'default':
                dob = ''
                data1['dob'] = None
            elif data.calendar_type == 'solar':
                data1['tithi'] = None
                data1['rasi'] = None
                data1['nakshatram'] = None
            else:
                data1['maasam'] = None
                data1['paksham'] = None
                data1['tithi'] = None

        if dob != '':
            data1['dob'] = datetime.strptime(dob, "%d %B, %Y")
        form = MemberForm(data1, instance=data)
        if form.is_valid():
            form.save()
            user = User.objects.get(pk=get_current_user().pk)
            if user.roles == "ROLE_MEMBER":
                return HttpResponseRedirect("%s?edit=1" %reverse('profile-view'))
            else:
                return HttpResponseRedirect("%s?edit=1" %reverse('member-view', kwargs={"mid": m_id}))
        return render(request, 'edit_member.html', locals())

@method_decorator(login_required, name="get")
class UserAdd(View):
    @method_decorator(has_permission_decorator('ACCESS_CONTROL'))
    def get(self, request):
        form = UserForm()
        member = Member.objects.all()
        return render(request, 'edit_user.html', locals())
   
    @method_decorator(has_permission_decorator('ACCESS_CONTROL'))
    def post(self, request):
        data1 = dict(request.POST.items())
        data1['created'] = datetime.now()
        if data1['username']:
           data1['username_canonical'] = data1['username'].strip().lower()
        if data1['email']:
           data1['email_canonical'] = self.normalize_email(data1['email'])
        data1['updated'] = datetime.now()
        data1['is_active'] = data1['enabled']
        data1['salt'] = get_random_string()
#        import ipdb; ipdb.set_trace()
        form = UserForm(data1)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect("%s?add=1" %reverse('user-access'))
        print form.errors
        return render(request, 'edit_user.html', locals())

    def normalize_email(self, email):
        # remove spaces at start and end of the and lowercase email address
        email = email.strip().lower()
        # split email into username and domain information
        username, domain = email.split('@')
        # remove . characters from username
        username = username.replace('.', '')
        #remove everything after +
        username = username.split('+')[0]

        return "%s@%s" % (username, domain)

@method_decorator(login_required, name="get")
class UserEdit(View):
    @method_decorator(has_permission_decorator('ACCESS_CONTROL'))
    def get(self, request, user_id):
        data = User.objects.get(pk=user_id)
        member = Member.objects.all()
        return render(request, 'edit_user.html', locals())
   
    @method_decorator(has_permission_decorator('ACCESS_CONTROL'))
    def post(self, request, user_id):
        data = User.objects.get(pk=user_id)
        data1 = dict(data.__dict__.items() + request.POST.items())
        data1['updated'] = datetime.now()
        data1['credentials_expire_at'] = datetime.strptime(request.POST.get('credentials_expire_at'), "%Y-%m-%d %H:%M")
        form = UserForm(data1, instance=data)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect("%s?edit=1" %reverse('user-access'))
        print form.errors
        return render(request, 'edit_user.html', locals())

@method_decorator(login_required, name="get")
class MemberAccess(View):
    @method_decorator(has_permission_decorator('ACCESS_CONTROL'))
    def get(self, request):
        users = User.objects.all()
        return render(request, 'member_access.html', locals())



@method_decorator(login_required, name="get")
class MemberAdd(View):
    @method_decorator(has_permission_decorator('ADD_MEMBER_DETAILS'))
    def get(self, request):
        form = MemberForm()
        maasam = MaasamType.objects.all()
        paksham = PakshamType.objects.all()
        tithi = TithiType.objects.all()
        rasi = RasiType.objects.all()
        nakshatrams = NakshatramType.objects.all()
        return render(request, 'edit_member.html', locals())

    @method_decorator(has_permission_decorator('ADD_MEMBER_DETAILS'))
    def post(self, request):
        data1 = dict(request.POST.items())
        data1['created'] = datetime.now()
        data1['updated'] = datetime.now()
        data1['created_by'] = get_current_user().username
        data1['updated_by'] = get_current_user().username
        data1['id'] = m_id = str(uuid.uuid4())
        dob = request.POST.get('dob')
        if dob != '':
            data1['dob'] = datetime.strptime(dob, "%d %B, %Y")
        from django.db.models import Max
        data1['mid'] = Member.objects.all().aggregate(Max('mid'))['mid__max']+1
        form = MemberForm(data1)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect("%s?add=1" %reverse('member-view', kwargs={"mid": m_id}))
        return render(request, 'edit_member.html', locals())


@method_decorator(login_required, name="get")
class SevaEditView(View):
    @method_decorator(has_permission_decorator('EDIT_MEMBER_SEVAS'))
    def get(self, request, s_id):
        sevaaddress = None
        data = Seva.objects.get(id=s_id)
        memdata = Member.objects.get(id=data.mid.id)
        form = SevaForm(instance=data)
        organisations = Organisation.objects.all()
        sevaCategories = SevaCategory.objects.all()
        maasam = MaasamType.objects.all()
        paksham = PakshamType.objects.all()
        tithi = TithiType.objects.all()
        raasi = RasiType.objects.all()
        nakshatrams = NakshatramType.objects.all()
        
        if data.inthenameof != 'donor' and data.inthenameof != 'Donor':
            try:
                sevaaddress = SevaAddress.objects.get(sid=s_id)
            except:
                pass
        return render(request, 'edit_seva.html', locals())

    @method_decorator(has_permission_decorator('EDIT_MEMBER_SEVAS'))
    def post(self, request, s_id):
        data = Seva.objects.get(id=s_id)
        inthenameof = data.inthenameof
        memdata = Member.objects.get(id=data.mid.id)
        data1 = dict(data.__dict__.items() + request.POST.items())
        data1['updated'] = datetime.now()
        data1['updated_by'] = get_current_user().username
        data1['mid'] = data.mid.id
        sevadate = request.POST.get('sevadatestd')

        if data.calendar_type != data1['calendar_type']:
            if data.calendar_type == 'default':
                sevadate = ''
                data1['sevadatestd'] = None
                data1['sevaday'] = ''
                data1['sevadate'] = None
                data1['startdate'] = None
            elif data.calendar_type == 'solar':
                data1['tithi'] = None
                data1['raasi'] = None
                data1['nakshatram'] = None
            else:
                data1['maasam'] = None
                data1['paksham'] = None
                data1['tithi'] = None
                 
        if sevadate != '':
            data1['startdate'] = datetime.strptime(sevadate, "%d %B, %Y")
            data1['sevaday'] = datetime.strptime(sevadate, "%d %B, %Y").strftime("%m%d")
            data1['sevadatestd'] = datetime.strptime(sevadate, "%d %B, %Y")
            data1['sevadate'] = datetime.strptime(sevadate, "%d %B, %Y")
            #data1['enddate']
        if data1['inthenameof'] != 'Donor' and data1['inthenameof'] != 'donor':
            data1['inthenameof'] = data1['name']
        form = SevaForm(data1, instance=data)
        if form.is_valid():
            form.save()
            sevaaddress = None
            if data1['inthenameof'] != 'Donor' and data1['inthenameof'] != 'donor':
                if inthenameof == 'Donor' or inthenameof == 'donor' :
                    sevaaddress = SevaAddress()
                    sevaaddress.sid = data
                else:
                    sevaaddress = SevaAddress.objects.get(sid=s_id)
                sevaaddress.address = data1['address']
                sevaaddress.city = data1['city']
                sevaaddress.district = data1['district']
                sevaaddress.state = data1['state']
                sevaaddress.country = data1['country']
                sevaaddress.phone = data1['phone']
                sevaaddress.isvalid = request.POST.get('isvalid', None)
                sevaaddress.pin = data1['pincode']
                sevaaddress.save()

            user = User.objects.get(pk=get_current_user().pk)
            if user.roles == "ROLE_MEMBER":
                return HttpResponseRedirect("%s?edit=2" %reverse('profile-view'))
            else:
                return HttpResponseRedirect("%s?edit=2" %reverse('member-view', kwargs={"mid": data.mid.id}))
        print form.errors
        return render(request, 'edit_seva.html', locals())

@method_decorator(login_required, name="get")
class SevaAdd(View):
    @method_decorator(has_permission_decorator('ADD_MEMBER_SEVAS'))
    def get(self, request, m_id=None):
        form = SevaForm()
        memdata = Member.objects.get(id=m_id)
        organisations = Organisation.objects.all()
        sevaCategories = SevaCategory.objects.all()
        maasam = MaasamType.objects.all()
        paksham = PakshamType.objects.all()
        tithi = TithiType.objects.all()
        raasi = RasiType.objects.all()
        nakshatrams = NakshatramType.objects.all()
        return render(request, 'edit_seva.html', locals())

    @method_decorator(has_permission_decorator('ADD_MEMBER_SEVAS'))
    def post(self, request, m_id=None):
        memdata = Member.objects.get(id=m_id)
        data1 = dict(request.POST.items())
        sevadate = request.POST.get('sevadatestd')
        if sevadate != '':
            data1['startdate'] = datetime.strptime(sevadate, "%d %B, %Y")
            data1['sevaday'] = datetime.strptime(sevadate, "%d %B, %Y").strftime("%m%d")
            data1['sevadatestd'] = datetime.strptime(sevadate, "%d %B, %Y")
            data1['sevadate'] = datetime.strptime(sevadate, "%d %B, %Y")
        data1['created'] = datetime.now()
        data1['updated'] = datetime.now()
        data1['created_by'] = get_current_user().username
        data1['updated_by'] = get_current_user().username
        data1['id'] = sid = str(uuid.uuid4())
        data1['mid'] = m_id
        if data1['inthenameof'] != 'Donor' and data1['inthenameof'] != 'donor':
            data1['inthenameof'] = data1['name']
        category = SevaCategory.objects.get(id=data1['sid'])
        data1['ssid']=  category.code+'#{}'.format(int(category.last_sequence_number)+1)
        category.last_sequence_number = str(int(category.last_sequence_number) + 1)
        category.save()
            
        form = SevaForm(data1)
        if form.is_valid():
            form.save()
            if data1['inthenameof'] != 'Donor' and data1['inthenameof'] != 'donor':
                seva = Seva.objects.get(id=sid)
                sevaaddress = SevaAddress()
                sevaaddress.id = str(uuid.uuid4())
                sevaaddress.sid = seva
                sevaaddress.address = data1['address']
                sevaaddress.city = data1['city']
                sevaaddress.district = data1['district']
                sevaaddress.state = data1['state']
                sevaaddress.country = data1['country']
                sevaaddress.phone = data1['phone']
                sevaaddress.pin = data1['pincode']
                sevaaddress.isvalid = False
                sevaaddress.save()
            user = User.objects.get(pk=get_current_user().pk)
            if user.roles == "ROLE_MEMBER":
                return HttpResponseRedirect("%s?add=2" %reverse('profile-view'))
            else:
                return HttpResponseRedirect("%s?add=2" %reverse('member-view', kwargs={"mid": m_id}))
        print form.errors
        return render(request, 'edit_seva.html', locals())

@method_decorator(login_required, name="get")
class MemberAddress(View):
    @method_decorator(has_permission_decorator('ADD_MEMBER_ADDRESS'))
    def get(self, request, m_id=None):
        address_list = Address.objects.filter(mid=m_id)
        return render(request, 'member/add_address.html', locals())

    @method_decorator(has_permission_decorator('ADD_MEMBER_ADDRESS'))
    def post(self, request, m_id=None):
        data1 = dict([('mid', m_id)] + request.POST.items())
        data1['created'] = datetime.now()
        data1['updated'] = datetime.now()
        data1['created_by'] = get_current_user().username
        data1['updated_by'] = get_current_user().username
        data1['id'] = str(uuid.uuid4())
        form = AddressForm(data1)
        if form.is_valid():
            form.save()
            user = User.objects.get(pk=get_current_user().pk)
            if user.roles == "ROLE_MEMBER":
                return HttpResponseRedirect("%s?add=3" %reverse('profile-view'))
            else:
                return HttpResponseRedirect("%s?add=3" %reverse('member-view', kwargs={"mid": m_id}))
        address_list = Address.objects.filter(mid=m_id)
        return render(request, 'member/add_address.html', locals())


@method_decorator(login_required, name="get")
class MemberAddressEdit(View):
    @method_decorator(has_permission_decorator('EDIT_MEMBER_ADDRESS'))
    def get(self, request, m_id=None, aid=None):
        address_list = Address.objects.filter(mid=m_id)
        data = Address.objects.get(id=aid)
        return render(request, 'member/edit_address.html', locals())

    @method_decorator(has_permission_decorator('EDIT_MEMBER_ADDRESS'))
    def post(self, request, m_id=None, aid=None):
        data = Address.objects.get(id=aid)
        data1 = dict(data.__dict__.items() + request.POST.items())
        data1['updated'] = datetime.now()
        data1['updated_by'] = get_current_user().username
        data1['mid'] = m_id
        data1['is_valid'] = True if request.POST.get('valid') else False
        data1['is_primary'] = True if request.POST.get('primary') else False
        form = AddressForm(data1, instance=data)
        if form.is_valid():
            form.save()
            user = User.objects.get(pk=get_current_user().pk)
            if user.roles == "ROLE_MEMBER":
                return HttpResponseRedirect("%s?edit=3" %reverse('profile-view'))
            else:
                return HttpResponseRedirect("%s?edit=3" %reverse('member-view', kwargs={"mid": m_id}))
        address_list = Address.objects.filter(mid=m_id)
        return render(request, 'member/edit_address.html', locals())

@method_decorator(login_required, name="post")
class MemberImageUpload(View):
   
    @method_decorator(has_permission_decorator('EDIT_MEMBER_DETAILS'))
    def post(self, request, m_id):
        data = Member.objects.get(id=m_id)
        f = request.FILES['file']
        from django.conf import settings
        file_path = settings.MEDIA_ROOT
        file_name = f.name
        extension =  file_name.split(".")[1]
        import os
        m_path = "%s/%s" % (file_path,"members")
        if not os.path.exists(file_path):
            os.makedirs(file_path)
            os.makedirs(m_path)
        elif not os.path.exists(m_path):
            os.makedirs(m_path)
        file_path = "%s/%s.%s" %(m_path, data.mid, extension)
        with open(file_path, 'wb+') as destination:
            for chunk in f.chunks():
                destination.write(chunk)
        data.photo = "/media/members/%s.%s"%(data.mid, extension)
        data.save()
        user = User.objects.get(pk=get_current_user().pk)
        if user.roles == "ROLE_MEMBER":
            return HttpResponseRedirect("%s?image=1" %reverse('profile-view'))
        else:
            return HttpResponseRedirect("%s?image=1" %reverse('member-view', kwargs={"mid": m_id}))

@method_decorator(login_required, name="get")
class OrganisationEdit(HasPermissionsMixin, View):
    required_permission = 'EDIT_ORGANISATION'

    def get(self, request, org_id=None):
        org_obj = Organisation.objects.get(id=org_id)
        form = OrganisationForm(instance=org_obj)
        return render(request, 'edit_organisation.html', locals())

    def post(self, request, org_id=None):
        org_obj = Organisation.objects.get(id=org_id)
        form = OrganisationForm(request.POST, instance=org_obj)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('organisation'))
        return render(request, 'edit_organisation.html', locals())


@method_decorator(login_required, name="get")
class ExportLocationReports(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
        loc = request.GET.get("loc", None)
        loc_list = Member.objects.filter(place__icontains=loc)
        response = HttpResponse(content_type='text/csv')
        cur_date = datetime.now().strftime('%d-%B-%Y')
        file_name = loc + "-" + cur_date + ".csv"
        response['Content-Disposition'] = "attachment; filename=%s"%file_name
     
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))
        writer.writerow([
            smart_str(u"Name"),
            smart_str(u"Place"),
            smart_str(u"Email"),
            smart_str(u"Phone")])
        for mem in loc_list:
            writer.writerow([
                smart_str(mem.name),
                smart_str(mem.place),
                smart_str(mem.email),
                smart_str(mem.phone)
            ])
     
        return response


class CheckMonthly:
    @staticmethod
    def is_monthly_report(request):
        month_input = request.GET.get("month_input", None)
        if month_input:
            return True
        else:
            return False

class CommonSevaSearch:
    @staticmethod
    def get_response(request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        cur_date = datetime.now().strftime('%d-%b-%Y')
        file_name = "seva" + "-" + cur_date + ".xlsx"
        response['Content-Disposition'] = 'attachment; filename=%s'%file_name
        wb = openpyxl.Workbook()
        wb.font = Font(name='Verdana')
        ws = wb.active
        ws.header_footer.center_header.font_name = "Verdana"
        ws.title = "Seva Report"
        sid = request.GET.get("sid", None)
        if CheckMonthly.is_monthly_report(request):
            month_input = request.GET.get("month_input", None)
            query_param = {'sid': sid, 'sevadate__icontains': "%s%s%s" % ("-", month_input, "-")}
        else:
            from_date = request.GET.get("from", None)
            to_date = request.GET.get("to", None)
            query_param = {'sid': sid, 'sevaday__range': [from_date, to_date]}
        seva_list = Seva.objects.filter(**query_param)
        row_num = 0
    
        columns = [
            (u"DONOR", 15),
            (u"MEMBER ID", 70),
            (u"SEVA ID", 70),
            (u"PLACE", 70),
            (u"IN THE NAME OF", 70),
            (u"SEVA DATE", 70),
            (u"ADDRESS", 70),
        ]
    
        for col_num in xrange(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            c.font = Font(bold=True, color="FF0000FF")
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
    
        for obj in seva_list:
            row_num += 1
            row = [
                obj.mid.name,
                "MID#%s" % obj.mid.mid,
                obj.ssid,
                obj.mid.place,
                obj.inthenameof,
                obj.sevadate,
                "%(address)s \n %(district)s \n %(state)s \n %(pin)s" % obj.mid.full_address.__dict__ if obj.mid.full_address else None,
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.style.alignment.wrap_text = True
    
        wb.save(response)
        return response

@method_decorator(login_required, name="get")
class ExportSevaSearch(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
        response = CommonSevaSearch.get_response(request)
        return response


@method_decorator(login_required, name="get")
class ExportMasamSevaSearch(View):

    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
        import openpyxl
        from openpyxl.cell import get_column_letter
        from openpyxl.styles import Font
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        cur_date = datetime.now().strftime('%d-%b-%Y')
        file_name = "sevamaasam" + "-" + cur_date + ".xlsx"
        response['Content-Disposition'] = 'attachment; filename=%s'%file_name
        wb = openpyxl.Workbook()
        wb.font = Font(name='Verdana')
        ws = wb.active
        ws.header_footer.center_header.font_name = "Verdana"
        ws.title = "Seva Maasam Report"
    
        sid = request.GET.get("sid", None)
        maasam_input = request.GET.get("maasam_input", None)
        query_param = {'sid': sid, 'maasam': maasam_input}
        seva_list = Seva.objects.filter(**query_param)
        
        row_num = 0
    
        columns = [
               (u"DONOR", 15),
               (u"MEMBER ID", 70),
               (u"SEVA ID", 70),
               (u"PLACE", 70),
               (u"IN THE NAME OF", 70),
               (u"SEVA TITHI", 70),
               (u"MAASAM", 70),
               (u"ADDRESS", 70),
            ]  
    
        for col_num in xrange(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            c.font = Font(bold=True, color="FF0000FF")
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
    
        for obj in seva_list:
            row_num += 1
            row = [
                obj.mid.name,
                "MID#%s" % obj.mid.mid,
                obj.ssid,
                obj.mid.place,
                obj.inthenameof,
                obj.tithi.tithi,
                obj.maasam.maasam,
                "%(address)s \n %(district)s \n %(state)s \n %(pin)s" % obj.mid.full_address.__dict__ if obj.mid.full_address else None,
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.style.alignment.wrap_text = True
    
        wb.save(response)
        return response


@method_decorator(login_required, name="get")
class ReportsMaasam(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
        sevas = SevaCategory.objects.all()
        maasams = MaasamType.objects.all()
        return render(request, 'reports_maasam.html', locals())

    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def post(self, request):
        sid = request.GET.get("sid", None)
        maasam_input = request.GET.get("maasam", None)
        query_param = {'sid': sid, 'maasam':maasam_input}
        seva_list = Seva.objects.filter(**query_param)
        return render(request, 'lunarssearch.html', locals())

@method_decorator(login_required, name="get")
class ExportAwardee(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
        import openpyxl
        from openpyxl.cell import get_column_letter
        from openpyxl.styles import Font
        awardee_list = list()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        cur_date = datetime.now().strftime('%d-%b-%Y')
        file_name = "Awardee" + "-" + cur_date + ".xlsx"
        response['Content-Disposition'] = 'attachment; filename=%s'%file_name
        wb = openpyxl.Workbook()
        wb.font = Font(name='Verdana')
        ws = wb.active
        ws.header_footer.center_header.font_name = "Verdana"
        ws.title = "Awardee Report"
    
        searchkey = request.GET.get("searchkey", None)

        search_by = request.GET.get("q", None)
        
        query_format = {}
        if search_by == "year":
                query_format['year__icontains'] = searchkey
        elif search_by == "specialization":
                query_format['specialization__icontains'] = searchkey
        elif search_by == "name":
                query_format['name__icontains'] = searchkey
        
        if query_format:
           awardee_list = Awardee.objects.filter(**query_format)
        else:
           awardee_list = Awardee.objects.all()
        
        row_num = 0
    
        columns = [
               (u"MEMBER ID", 70),
               (u"MEMBER NAME", 70),
               (u"AWARD", 70),
               (u"YEAR", 70),
               (u"SPECIALIZATION", 70),
               (u"PLACE", 70),
             
            ]  
    
        for col_num in xrange(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            c.font = Font(bold=True, color="FF0000FF")
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
    
        for obj in awardee_list:
            row_num += 1
            row = [
                "MID#%s" % obj.member.mid,
                obj.member.name,
                obj.name,
                obj.year,
                obj.specialization,
                obj.member.place,
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.style.alignment.wrap_text = True
    
        wb.save(response)
        return response

@login_required
@method_decorator(has_permission_decorator('EDIT_MEMBER_DETAILS'))
def edit_member(request, m_id=None):
    required_permission = 'ADD_ORGANISATION'
    data = Member.objects.get(id=m_id)
    if request.method == "POST":
        data1 = dict(data.__dict__.items() + request.POST.items())
        form = MemberForm(data1, instance=data)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('member-view', kwargs={"mid": m_id}))
    else:
        form = MemberForm(instance=data)
    maasam = MaasamType.objects.all()
    paksham = PakshamType.objects.all()
    tithi = TithiType.objects.all()
    rasi = RasiType.objects.all()
    nakshatrams = NakshatramType.objects.all()
    return render(request, 'edit_member.html', locals())


@login_required
@method_decorator(has_permission_decorator('ADD_MEMBER_DETAILS'))
def add_member(request):
    if request.method == "POST":
        data1 = dict(request.POST.items())
        data1['created'] = datetime.now()
        data1['updated'] = datetime.now()
        data1['id'] = m_id = str(uuid.uuid4())
        from django.db.models import Max
        data1['mid'] = Member.objects.all().aggregate(Max('mid'))['mid__max']+1
        form = MemberForm(data1)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('member-view', kwargs={"mid": m_id}))
        else:
            print form.errors
    else:
        form = MemberForm()
    maasam = MaasamType.objects.all()
    paksham = PakshamType.objects.all()
    tithi = TithiType.objects.all()
    rasi = RasiType.objects.all() 
    nakshatrams = NakshatramType.objects.all()
    return render(request, 'edit_member.html', locals())

@method_decorator(login_required, name="get")
class GroupView(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
        groups = GroupName.objects.all()
         
        return render(request, 'group_view.html', locals())

@method_decorator(login_required, name="get")
class ImportGroup(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request):
         
        return render(request, 'import_group.html', locals())

    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def post(self, request):
        filename = request.FILES.get('file')
        data1 = dict(request.POST.items())
        group = data1['group']
        file_read= filename.name
        file_type= file_read.split(".")[-1]
        newgroup = None

        if file_type == 'xlsx' or file_type == 'xls' or file_type == 'csv':
            try:
                newgroup = GroupName()
                newgroup.name = group
                newgroup.id =  gid =  str(uuid.uuid4())
                newgroup.save()
            except:
                return HttpResponseRedirect("%s?gname=1" %reverse('import-group'))
        else:
            print "Please upload Csv/Xlsx/xls files"
            return render(request, 'import_group.html', locals())
   
        form = ImportGroupForm(request.POST, request.FILES)
        if form.is_valid():
            if file_type == 'xlsx' or file_type == 'xls':
                self.importxls(filename, newgroup)
            elif file_type == 'csv':
                self.importcsv(filename, newgroup)
                
            return HttpResponseRedirect("%s?add=1" %reverse('group-view'))
        print form.errors
        return render(request, 'import_group.html', locals())


    def importxls(self, filename, newgroup):
        wb = openpyxl.load_workbook(filename, use_iterators=True)
        sheet = wb.active

        for row in range(1, sheet.max_row + 1):
            input_data = GroupData()
            input_data.name  = sheet['A' + str(row)].value
            input_data.surname  = sheet['B' + str(row)].value
            input_data.email = sheet['C' + str(row)].value
            input_data.mobile    = sheet['D' + str(row)].value
            input_data.gid = newgroup
            input_data.save()
       
    def importcsv(self, filename, newgroup):
        records = csv.reader(filename)

        for line in records:
            input_data = GroupData()
            input_data.name = line[0]
            input_data.gid = newgroup
            input_data.surname = line[1]
            input_data.email =   line[2]
            input_data.mobile = line[3]
            input_data.save()


@method_decorator(login_required, name="get")
#from django.contrib.contenttypes.models import ContentType
class MedicalProfileView(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request, m_id):
        try:
            mProfile = MedicalProfile.objects.get(mid=m_id)
            return HttpResponseRedirect(reverse('admin:member_medicalprofile_change', args=(mProfile.id,)))
        except Exception, e:
            return HttpResponseRedirect(reverse('admin:member_medicalprofile_add'))#, args=(mid=m_id)))

class MembershipAdd(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request, m_id):
        membership = request.GET.get("membership", None)
        trustee = Trustee.objects.filter(member=m_id)
        staffprofile = None
        honorary = Honorary.objects.filter(member=m_id)
        mem = Member.objects.get(pk=m_id)
        assetBuildings = AssetBuilding.objects.filter(maindonors=m_id)
        assetLands = AssetLand.objects.filter(inchargeperson=m_id)
        assetMoney =  DonationCash.objects.filter(member=m_id)
        assetEquipments = AssetEquipment.objects.filter(maindonors=m_id)
        return render(request, "add_membership.html", locals())

    @method_decorator(has_permission_decorator('EDIT_MEMBER_DETAILS'))
    def post(self, request, m_id):
        membership = request.POST.get("membership", None)
        if membership == 'money':
            return HttpResponseRedirect(reverse('admin:member_donationcash_add'))
        elif membership == 'land':
            return HttpResponseRedirect(reverse('admin:member_assetland_add'))
        elif membership == 'building':
            return HttpResponseRedirect(reverse('admin:member_assetbuilding_add'))
        elif membership == 'equipment':
            return HttpResponseRedirect(reverse('admin:member_assetequipment_add'))
        elif membership == 'trustee':
            return HttpResponseRedirect(reverse('admin:member_trustee_add'))
        elif membership == 'inmate':
            return HttpResponseRedirect(reverse('admin:member_staffprofile_add'))
        elif membership == 'medical':
            return HttpResponseRedirect(reverse('admin:member_medicalprofile_add'))
        elif membership == 'honorary':
            return HttpResponseRedirect(reverse('admin:member_honorary_add'))
        elif membership == 'complimentary':
            return HttpResponseRedirect(reverse('admin:member_complimentary_add'))

@method_decorator(login_required, name="get")
class MembershipView(View):
    required_permission = 'VIEW_MEMBER_DETAILS'

    def get(self, request, m_id):
        trustee = Trustee.objects.filter(member=m_id)
        staffprofile = None
        honorary = Honorary.objects.filter(member=m_id)
        #mem = Member.objects.get(pk=mid)
        assetBuildings = AssetBuilding.objects.filter(maindonors=m_id)
        assetLands = AssetLand.objects.filter(inchargeperson=m_id)
        assetMoney = DonationCash.objects.filter(member=m_id)
        assetEquipments = AssetEquipment.objects.filter(maindonors=m_id)
        return render(request, "add_membership.html", locals())

@method_decorator(login_required, name="get")
class AddRelationship(View):
    @method_decorator(has_permission_decorator('EXPORT_REPORTS'))
    def get(self, request, m_id):
        return render(request, "add_relationship.html", locals())
    
    def post(self, request, m_id):
        relationship = request.POST.get("relationship", None)
        #data1 = dict(request.POST.items())
        relation = None
        try:
            relation = Relatives.object.get(mid=m_id)
        except:
            relation = Relatives()
            relation.mid=m_id
        ''' if relationship == 'father':
            reltaion.father = 
        elif relationship == 'mother':
            relation.mother = 
        elif relationship == 'friend':
            relation.friend =
        elif relationship == 'wife':
            realation.wife = 
        elif relationship == 'husband':
            relationship.husband = 
        elif relationship == 'sister':
            relationship.sister =
        elif relationship == 'brother':
            relationship.brother = 
        elif relationship == 'son':
            relationship.son =
        elif relationship == 'daughter':
            relationship.daughter =
        elif relationship == 'cousins':
            relationship.cousins =
        elif relationship == 'inlaws':
            relationship =
        elif relationship == 'wellwisher':
            relationship =
         elif relationship == 'guru':
            relationship =   '''
       
        relation.save()

        return HttpResponseRedirect(reverse('member-view', kwargs={"mid": m_id}))
        

@method_decorator(login_required, name="get")
class RelationSearchView(HasPermissionsMixin, View):
    required_permission = 'VIEW_MEMBER_DETAILS'
   
    def get(self, request):
        search_by = request.GET.get("q", None)
        data = request.GET
	print data
        query_format = {}
            
        print search_by
        if search_by == 'surname':
            query_format['surname__icontains'] = data["surname"]
        elif search_by == "mid":
            if data.get('mid'):
                query_format['mid__exact'] = data["mid"]
        if query_format:
            print query_format 
            result = Member.objects.filter(**query_format)

        return render(request, 'rsearch.html', locals())
